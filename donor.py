import consts
import validation
import openpyxl as op


def donation():
    wb = op.load_workbook(r'gear_amounts.xlsx')
    gear = input("what do you want to donate?")
    if gear in consts.MIN_GEAR_DICT:
        bases_lack = []
        for base in consts.BASE_CODES_DICT:
            ws = wb[base]
            row_gear = check_row(gear)
            if ws.cell(row=row_gear, column=consts.AMOUNT_COL).value < consts.MIN_GEAR_DICT[gear][base]:
                bases_lack.append(base)
        print("the red ones are the most urgent")
        for base in bases_lack:
            ws = wb[base]
            row_gear = check_row(gear)
            if ws.cell(row=row_gear, column=consts.URGENCY_COL).value == "Urgent" :
                print("red")
            else:
                print("black")
        chosen_base = input("which base do you choose?")
        if chosen_base in bases_lack:
            print("enter your details")
            amount = int(input("how many do you want to donate?"))
            gmail = input("what is your email?")
            if validation.checking_mail(gmail):
                pass# send the donor details
        return bases_lack

def check_row(gear):
    if gear == "Vests":
        row = consts.VESTS_ROW
    elif gear == "Gloves":
        row = consts.GLOVES_ROW
    elif gear == "Helmets":
        row = consts.HELMETS_ROW
    elif gear == "Goggles":
        row = consts.GOGGLES_ROW
    elif gear == "Food":
        row = consts.FOOD_ROW
    return row

