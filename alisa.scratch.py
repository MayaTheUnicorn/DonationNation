import openpyxl as op
import consts
import tkinter as tk
import consts
from tkinter import OptionMenu, StringVar


# wb = op.load_workbook(r'gear_amounts.xlsx')
# ws = wb['Uvda']
# print(ws['B2'].value)
# ws['B2'].value = 5
# wb.save(r'C:\Users\kotbe\OneDrive\Desktop\Donation_Nation.xlsx')
# # print(ws.cell(row=2, column=3).value)
# import tkinter as tk
# from tkinter import OptionMenu, StringVar
#
#

def add_logo():
   confetti = tk.PhotoImage(file="confetti.gif")  # loading image
   confetti_label = tk.Label(soldier_screen_action, confetti)  # creating label element
   # confetti_label.place(x=150, y=350)  # placement
   confetti_label.pack()




def done():
   print("yay")


soldier_screen_action = tk.Tk()
soldier_screen_action.geometry("400x400")
soldier_screen_action.configure(bg=consts.BACKGROUND_COLOR)


clicked = StringVar()
clicked.set("choose an action")


drop = OptionMenu(soldier_screen_action, clicked, "Add gear", "Remove gear", "Urgent/not Urgent")
drop.pack()


button = tk.Button(soldier_screen_action, text="done", command=done)

add_logo()
soldier_screen_action.mainloop()
#
#
#
#
#
#
# def soldier_action(logged_in):
#     if logged_in:
#         action = input("choose an action: 1- add gear, 2- remove gear, 3- urgent/not urgent")
#         if action == 1:
#             gear_object = input("what gear do you want to add?")
#             if not gear_object in consts.MIN_GEAR_DICT:
#                 gear_object = input("what gear do you want to add?")
#
#             gear_count = input("how much gear do you want to add?")
#
#             wb = op.load_workbook(r'gear_amounts.xlsx')
#
#             for base in consts.BASE_LIST:
#                 ws = wb[base]
            # print(ws['B2'].value)
            # ws['B2'].value = 5
            # wb.save(r'C:\Users\kotbe\OneDrive\Desktop\Donation_Nation.xlsx')
            # # print(ws.cell(row=2, column=3).value)
            # import tkinter as tk
            # from tkinter import OptionMenu, StringVar

