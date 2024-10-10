import tkinter as tk
import consts
import openpyxl as op

import donor
import screen


# Check if soldier entered correct password
def soldier_login(list):
    if consts.BASE_CODES_DICT[list[0]] == list[1]:
        return soldier_action(list[0])
    else:
        return False

def soldier_action(base):

    action = input("choose an action: 1- add gear, 2- remove gear")
    if action == 1:
        gear_object = input("what gear do you want to add?")
        if gear_object in consts.MIN_GEAR_DICT:
            gear_amount = input("how much gear do you want to add?")
            wb = op.load_workbook(r'gear_amounts.xlsx')
            ws = wb[base]
            row_gear = donor.check_row(gear_object)
            ws.cell(row=row_gear, column=consts.AMOUNT_COL).value += gear_amount
            wb.save(r'gear_amounts.xlsx')

    if action == 2:
        gear_object = input("what gear do you want to remove?")
        if gear_object in consts.MIN_GEAR_DICT:
            gear_amount = input("how much gear do you want to remove?")
            wb = op.load_workbook(r'gear_amounts.xlsx')
            ws = wb[base]
            row_gear = donor.check_row(gear_object)
            ws.cell(row=row_gear, column=consts.AMOUNT_COL).value -= gear_amount
            wb.save(r'gear_amounts.xlsx')


